"""File upload API for chat — saves files to agent workspace and extracts text."""

import base64
import os
import uuid
from pathlib import Path
from tempfile import NamedTemporaryFile

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, Form
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.permissions import check_agent_access
from app.core.security import get_current_user
from app.database import get_db
from app.models.user import User
from app.services.media_assets import (
    MediaContractError,
    image_asset_from_bytes,
    validate_uploaded_video,
)
from app.services.storage import ensure_local_path, get_storage_backend, guess_content_type, normalize_storage_key

router = APIRouter(prefix="/chat", tags=["chat"])

# Supported extensions and their text extraction method
TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".xml", ".yaml", ".yml",
    ".py", ".js", ".ts", ".html", ".css", ".sql", ".sh", ".log",
    ".ini", ".cfg", ".conf", ".env", ".toml",
}
OFFICE_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv"}
EXTRACTABLE = TEXT_EXTENSIONS | OFFICE_EXTENSIONS

MIME_MAP = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".mp4": "video/mp4", ".avi": "video/x-msvideo", ".mov": "video/quicktime",
    ".mkv": "video/x-matroska",
}

MAX_CHAT_UPLOAD_BYTES = 25 * 1024 * 1024


def _validate_multimodal_upload_extension(extension: str) -> None:
    """Reject formats that local decoders accept but MiniMax M3 does not."""

    if extension in {".bmp", ".gif"}:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{extension.removeprefix('.').upper()} is not supported for multimodal chat. "
                "Convert the image to JPEG, PNG, or WEBP before uploading."
            ),
        )


async def _read_upload_with_limit(file: UploadFile, limit: int = MAX_CHAT_UPLOAD_BYTES) -> bytes:
    """Read an upload without allowing an unbounded request body in memory."""
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(min(1024 * 1024, limit + 1 - total))
        if not chunk:
            break
        chunks.append(chunk)
        total += len(chunk)
        if total > limit:
            raise HTTPException(status_code=413, detail="File too large (max 25MB)")
    return b"".join(chunks)


def extract_text(file_path: Path, extension: str) -> str:
    """Extract text content from a file."""
    if extension in TEXT_EXTENSIONS:
        try:
            return file_path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return file_path.read_text(encoding="gbk", errors="replace")

    if extension == ".pdf":
        try:
            import fitz

            with fitz.open(file_path) as document:
                text = "\n".join(page.get_text() or "" for page in document)
            return text[:8000].strip() or "[PDF内容提取失败]"
        except Exception as e:
            return f"[PDF解析错误: {e}]"

    if extension == ".docx":
        try:
            from docx import Document

            document = Document(str(file_path))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return text[:8000].strip() or "[DOCX内容提取失败]"
        except Exception as e:
            return f"[DOCX解析错误: {e}]"

    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, read_only=True, data_only=True)
            lines: list[str] = []
            for worksheet in workbook.worksheets[:3]:
                lines.append(f"## Sheet: {worksheet.title}")
                for row in worksheet.iter_rows(max_row=50, values_only=True):
                    lines.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            workbook.close()
            return "\n".join(lines)[:8000].strip() or "[Excel内容提取失败]"
        except Exception as e:
            return f"[Excel解析错误: {e}]"

    return f"[不支持的文件格式: {extension}]"


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    agent_id: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload a file for chat context. Saves to agent workspace/uploads/ and returns extracted text."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")

    ext = os.path.splitext(file.filename)[1].lower()
    _validate_multimodal_upload_extension(ext)

    content = await _read_upload_with_limit(file)
    is_image = ext in IMAGE_EXTENSIONS
    is_video = ext in VIDEO_EXTENSIONS
    if is_image and len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Image too large (max 10MB)")
    try:
        if is_image:
            image_asset_from_bytes(content, label="Uploaded image")
        elif is_video:
            await validate_uploaded_video(
                content,
                extension=ext,
                label="Uploaded video",
            )
    except MediaContractError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    # Determine save directory
    workspace_path = ""
    saved_filename = ""
    if agent_id:
        try:
            parsed_agent_id = uuid.UUID(agent_id)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Invalid agent_id") from exc
        await check_agent_access(db, current_user, parsed_agent_id)
        storage = get_storage_backend()
        filename = file.filename.replace("/", "_").replace("\\", "_")
        stem, suffix = os.path.splitext(filename)
        # Object stores and local storage both overwrite an existing key.  A
        # check-then-write loop is therefore racy when two users upload the
        # same name concurrently.  Allocate a unique key up front instead.
        stored_filename = f"{stem}_{uuid.uuid4().hex[:12]}{suffix}"
        saved_filename = stored_filename
        workspace_path = f"workspace/uploads/{stored_filename}"
        key = normalize_storage_key(f"{parsed_agent_id}/{workspace_path}")
        await storage.write_bytes(
            key,
            content,
            content_type=guess_content_type(stored_filename),
        )
        save_path = await ensure_local_path(key)
    else:
        # Fallback: save to /tmp (legacy behavior)
        fallback_dir = Path("/tmp/clawith_uploads")
        fallback_dir.mkdir(mode=0o700, exist_ok=True)
        suffix = Path(file.filename.replace("\\", "/")).suffix[:20]
        with NamedTemporaryFile(
            mode="wb",
            prefix="chat-upload-",
            suffix=suffix,
            dir=fallback_dir,
            delete=False,
        ) as handle:
            handle.write(content)
            save_path = Path(handle.name)
        saved_filename = save_path.name

    # Extract text (only for known formats)
    image_data_url = ""
    video_data_url = ""
    if is_image:
        # For images: generate base64 data URL for vision models
        mime = MIME_MAP.get(ext, "image/png")
        b64 = base64.b64encode(content).decode("ascii")
        image_data_url = f"data:{mime};base64,{b64}"
        extracted = f"[图片文件: {file.filename}，需要视觉模型分析]"
    elif is_video:
        mime = MIME_MAP.get(ext, "video/mp4")
        b64 = base64.b64encode(content).decode("ascii")
        video_data_url = f"data:{mime};base64,{b64}"
        extracted = f"[视频文件: {file.filename}，需要视频理解模型分析]"
    elif ext in EXTRACTABLE:
        extracted = extract_text(save_path, ext)
    else:
        extracted = f"[文件已保存，格式 {ext} 暂不支持文本提取，Agent 可通过 read_document 工具读取]"

    # Truncate if too long
    if len(extracted) > 6000:
        extracted = extracted[:6000] + "\n\n...[内容已截断，共 " + str(len(extracted)) + " 字]"

    return {
        "filename": file.filename,
        "saved_filename": saved_filename,
        "size": len(content),
        "extracted_text": extracted,
        "workspace_path": workspace_path,
        "is_image": is_image,
        "is_video": is_video,
        "image_data_url": image_data_url,
        "video_data_url": video_data_url,
    }
